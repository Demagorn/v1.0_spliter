import time
import datetime as dt
import openpyxl.worksheet.copier
import pandas as pd
from openpyxl import load_workbook
from openpyxl import workbook, Workbook
from openpyxl import utils
from openpyxl.utils.dataframe import dataframe_to_rows
from copy import copy
import pathlib
from openpyxl.styles.protection import Protection
from notebook import insert_formulas,comments,lock_sheet

EXPORT_FOLDER_PATH = "./Export"
FORMATE = "Format.xlsx"
DATA_FILE = "1.xlsx"
COLORS_LIST = ["00FFFF00","00FF00FF","0000FFFF",
"00FF0000","0000FF00","000000FF","00FFFF00","00FF00FF",
"0000FFFF","00800000","00008000","00000080","00808000",
"00800080","00008080","00C0C0C0","00808080","009999FF"]
COLUMN_INDEX = "Employee #"
now = dt.datetime.now()
curent_time = now.strftime("%Y_%m_%d_%H_%M")


def create_df():
    """creates df with headers and formulas as string"""
    wb_data = load_workbook("1.xlsx")
    ws_data = wb_data.active
    df2 = pd.read_excel("1.xlsx")
    headers = df2.iloc[9].values
    df = pd.DataFrame(ws_data.values)
    # the drop is because the panda recognize empty column but th openpyxl not/ so i drop i manualy
    df = df.drop(columns=103)
    df.columns = headers
    df = df[11:]   #my main data starts on row 12
    df = df.reset_index(drop=True)
    return df


def create_beauty(amount_of_rows,amount_of_columns,worksheet):
    """makes a same cell format. to a data only"""
    for row in range(12,amount_of_rows+13):
        for cell in range(1,amount_of_columns):
            source_cell = worksheet.cell(column=cell, row=12)
            target_cell = worksheet.cell(column=cell, row=row)
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)

            if source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)

            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)
    worksheet.delete_rows(12)


def create_file_director(name,list_of_managers):
    """Creates directory and a file according to a format for director"""
    file_folder = f"Export/{name}"
    pathlib.Path(file_folder).mkdir(parents=True, exist_ok=True)
    wb = load_workbook(FORMATE)
    ws = wb["format"]
    for manager in list_of_managers:
        ws2 = wb.copy_worksheet(ws)
        ws2.title = manager
        ws2.sheet_properties.tabColor = COLORS_LIST[list_of_managers.index(manager)]
        ws2.sheet_view.showGridLines = False
    del wb["format"]
    wb.save(f"{file_folder}/{name}.xlsx")
    return file_folder

def create_file_manager(folder,name):
    """Creates directory and a file according to a format for manager"""
    file_folder = f"{folder}/{name}"
    pathlib.Path(file_folder).mkdir(parents=True, exist_ok=True)
    wb = load_workbook(FORMATE)
    ws = wb["format"]
    ws2 = wb.copy_worksheet(ws)
    ws2.title = name
    ws2.sheet_properties.tabColor = COLORS_LIST[1]
    ws2.sheet_view.showGridLines = False
    del wb["format"]
    wb.save(f"{file_folder}/{name}.xlsx")
    return file_folder

def lock_comment_formulas(target_file, source_file):
    insert_formulas(target_file)
    comments(source_file, target_file)
    lock_sheet(target_file)


start_time = time.perf_counter()
df = create_df()
director_list = list(df["Director"].unique())
for director in director_list[0:2]:
    df_director = df.loc[df["Director"] == director]
    group_managers = list(df_director["Group Manager"].unique())
    director_folder = create_file_director(director,group_managers)
    print("created")
    director_file = f"{director_folder}/{director}.xlsx"
    wb = load_workbook(director_file)
    for manager in group_managers:
        print(f"Starting manager {manager}")
        ws_director = wb[manager]
        manager_folder = create_file_manager(director_folder,manager)
        manager_file = f"{manager_folder}/{manager}.xlsx"
        wb_manager = load_workbook(manager_file)
        ws_manager = wb_manager[manager]
        df_manager=df_director.loc[df_director["Group Manager"]==manager]
        for r in dataframe_to_rows(df_manager, index=False, header=False):
            ws_director.append(r)
            ws_manager.append(r)
        create_beauty(df_manager["Group Manager"].size, len(df_director.columns), ws_director)
        create_beauty(df_manager["Group Manager"].size, len(df_director.columns),ws_manager)
        wb_manager.save(manager_file)
        lock_comment_formulas(manager_file,DATA_FILE)
    wb.save(director_file)
    director_time = time.perf_counter()
    print(f"Time to director {director} --- %s seconds ---"% (director_time - start_time))
    lock_comment_formulas(director_file,DATA_FILE)
end_time = time.perf_counter()
print("TOTAL TIME --- %s seconds ---" % (end_time - start_time))
