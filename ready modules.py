import time
import datetime as dt

import openpyxl.worksheet.copier
import pandas as pd
from openpyxl import load_workbook
from openpyxl import workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from copy import copy
import pathlib

EXPORT_FOLDER_PATH = "./Export"
FORMATE = "Format2.xlsx"
DATA_FILE = "1.xlsx"
COLORS_LIST = ["00FFFF00","00FF00FF","0000FFFF",
"00FF0000","0000FF00","000000FF","00FFFF00","00FF00FF",
"0000FFFF","00800000","00008000","00000080","00808000",
"00800080","00008080","00C0C0C0","00808080","009999FF"]
now = dt.datetime.now()
curent_time = now.strftime("%Y_%m_%d_%H_%M")

# def add_protection(worksheet,columns_open,range_of_cells):
#     worksheet.protection.sheet = True
#     for column in columns_open:
#         for cell in range_of_cells:
#             worksheet.cell(row=cell,column=column).protection = False


def create_beauty(amount_of_rows,amount_of_columns,worksheet):
    for row in range(12,amount_of_rows+13):
        for cell in range(1,amount_of_columns):
            source_cell = worksheet.cell(column=cell, row=12)
            target_cell = worksheet.cell(column=cell, row=row)
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)

            if source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)

            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)

    worksheet.delete_rows(12)




def create_file_director(name,list_of_managers):
    # Creates directory and a file according to a format for director
    file_folder = f"Export/{name}"
    pathlib.Path(file_folder).mkdir(parents=True, exist_ok=True)
    wb = load_workbook(FORMATE)
    ws = wb["format"]
    for manager in list_of_managers:
        ws2 = wb.copy_worksheet(ws)
        ws2.title = manager
        ws2.sheet_properties.tabColor = COLORS_LIST[list_of_managers.index(manager)]
        ws2.sheet_view.showGridLines = False
    del wb["format"]
    wb.save(f"{file_folder}/{name}.xlsx")
    return file_folder


def create_file_manager(folder,name):
    # Creates directory and a file according to a format for manager
    file_folder = f"{folder}/{name}"
    pathlib.Path(file_folder).mkdir(parents=True, exist_ok=True)
    wb = load_workbook(FORMATE)
    ws = wb["format"]
    ws2 = wb.copy_worksheet(ws)
    ws2.title = name
    ws2.sheet_properties.tabColor = COLORS_LIST[1]
    ws2.sheet_view.showGridLines = False
    del wb["format"]
    wb.save(f"{file_folder}/{name}.xlsx")
    return file_folder


start_time = time.time()
df = pd.read_excel(DATA_FILE,header=10)
director_list = list(df["Director"].unique())
for director in director_list[:1]:
    df_director = df.loc[df["Director"] == director]
    group_managers = list(df_director["Group Manager"].unique())
    director_folder = create_file_director(director,group_managers)
    print("created")
    # wb = load_workbook(f"{director_folder}/{director}.xlsx")
    # for manager in group_managers[:3]:
    #     ws_director = wb[manager]
    #     manager_folder = create_file_manager(director_folder,manager)
    #     wb_manager = load_workbook(f"{manager_folder}/{manager}.xlsx")
    #     ws_manager = wb_manager[manager]
    #     df_manager=df_director.loc[df_director["Group Manager"]==manager]
    #     for r in dataframe_to_rows(df_manager, index=False, header=False):
    #         ws_director.append(r)
    #         ws_manager.append(r)
    #     create_beauty(df_manager["Group Manager"].size, df_director.iloc[12].size, ws_director)
    #     create_beauty(df_manager["Group Manager"].size, df_director.iloc[12].size,ws_manager)
    #     wb_manager.save(f"{manager_folder}/{manager}.xlsx")
    # wb.save(f"{director_folder}/{director}.xlsx")
    print("Time to director --- %s seconds ---" % (time.time() - start_time))
print("TOTAL TIME --- %s seconds ---" % (time.time() - start_time))

