import time
import datetime as dt
import openpyxl.utils
import pandas as pd
from openpyxl import load_workbook
from openpyxl import workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from copy import copy
from openpyxl.styles.protection import Protection
import pathlib
import os
COLORS_LIST = ["00FFFF00","00FF00FF","0000FFFF",
"00FF0000","0000FF00","000000FF","00FFFF00","00FF00FF",
"0000FFFF","00800000","00008000","00000080","00808000",
"00800080","00008080","00C0C0C0","00808080","009999FF"]
FORMATE = "Format.xlsx"
EXPORT_FOLDER_PATH = "./Export"
DATA_FILE = "1.xlsx"

def create_root_file():
    """Creates directory and a file according to a format for director"""
    wb = load_workbook(FORMATE)
    ws = wb["format"]
    ws2 = wb.copy_worksheet(ws)
    ws2.title = "DATA"
    ws2.sheet_view.showGridLines = False
    del wb["format"]
    wb.save("2.xlsx")

def create_df():
    """ creates df with headers and formulas as string"""
    wb_data = load_workbook("1.xlsx")
    ws_data = wb_data.active
    df2 = pd.read_excel("1.xlsx")
    headers = df2.iloc[9].values
    df = pd.DataFrame(ws_data.values)
    df = df.drop(columns=103)
    df.columns = headers
    df = df[11:]
    df = df.reset_index(drop=True)
    return df

def locate_column(target_worksheet,header):
    for column, cellObj in enumerate(target_worksheet[11]):
        if cellObj.value == header: return column

def comments(source_file,target_file):
    """find all comments in ws and assign them to new ws"""
    wb = load_workbook(source_file)
    ws =wb.active
    comments_dic = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.comment:
                key = row[2].value
                comments_dic.setdefault(key,[])
                comments_dic[key].append([ws.cell(column=cell.column,row=11).value , cell.comment])
    wb2= load_workbook(target_file)
    sheets = wb2.sheetnames
    for sheet in sheets[3:]:
        ws2 = wb2[sheet]
        for index,row in enumerate(ws2.iter_rows()):
            if row[2].value in comments_dic:
                worker_id = row[2].value
                for comment in comments_dic[worker_id]:
                    column = locate_column(ws2,comment[0])
                    ws2.cell(column=column+1,row=index+1).comment = comment[1]
    wb2.save(target_file)


def comments_for_data_update(source_file,target_file):
    """find all comments in ws and assign them to new ws"""
    wb = load_workbook(source_file)
    comments_dic = {}
    sheets = wb.sheetnames
    for sheet in sheets:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment:
                    key = row[2].value
                    comments_dic.setdefault(key,[])
                    comments_dic[key].append([ws.cell(column=cell.column,row=11).value , cell.comment])
    wb2= load_workbook(target_file)
    ws2 = wb2["DATA"]
    comment_writen = {}
    for key in comments_dic:
        for index, row in enumerate(ws2.iter_rows()):
            if key == row[2].value:
                for comment in comments_dic[key]:
                    column = locate_column(ws2,comment[0])
                    ws2.cell(column=column+1,row=index+1).comment = comment[1]
                    comment_writen.setdefault(key, [])
                    comment_writen[key].append([column , comment[1]])
    wb2.save(target_file)



def write_on_every_line(worksheet,text, column):
    """adding formulas for every row:"""
    for row, cellObj in enumerate(list(worksheet.columns)[column]):
        if row >= 11:
            n = text.format(row+1)
            # n = '=(E%d-F%d)/C%d' % (row+1, row+1, row+1)
            cellObj.value = n


def create_formulas():
    """create formulas for insertion"""
    wb= load_workbook("Roberto Pozzi Southern Europe_2021-CompReview_APPROVED_SAMPLE.xlsx")
    ws = wb["Southern Europe Mgmt"]
    for row in ws.iter_rows(min_row=13,max_row=13):
        formulas_dic = {}
        for column, cellObj in enumerate(row):
            if cellObj.data_type == "f" : formulas_dic [ws[11][column].value] =cellObj.value
        for key, value in formulas_dic.items():
            formulas_dic[key] = value.replace("13","{0}")
        return formulas_dic

def insert_formulas(target_file):
    """adds formulas to file"""
    wb = load_workbook(target_file)
    sheets = wb.sheetnames
    for sheet in sheets[3:]:
        formulas_dic = create_formulas()
        for key,value in formulas_dic.items():
            ws = wb[sheet]
            column = locate_column(ws,key)
            write_on_every_line(ws,value,column)
    wb.save(target_file)


def lock_sheet(target_file):
    open_columns = []
    wb = load_workbook("Format.xlsx")
    ws = wb["format"]
    for column,cell in enumerate(ws[12]):
        if cell.fill.start_color.index == 0:
            open_columns.append(column)
    wb2 = load_workbook(target_file)
    sheets = wb2.sheetnames
    for sheet in sheets:
        ws2 = wb2[sheet]
        ws2.protection.sheet = True
        for index in open_columns:
            column = openpyxl.utils.get_column_letter(index+1)
            for cell in ws[column]:
                if cell.row >= 12:
                    cell.protection = Protection(locked=False)
    wb2.save(target_file)





def add_filter(ws):
    """adds filter to a headers row""" #need to fix amout of columns and rows acoornid to a df size
    fullrange = "B11:" + get_column_letter(ws.max_column) \
                + str(ws.max_row)
    ws.auto_filter.ref = fullrange


COLUMN_INDEX = "Employee #"


def update_main_data(target_file):
    files_for_update = os.listdir("./Files for update")
    df_root = pd.read_excel(target_file, header=10)
    df_header_list = df_root.columns.to_list()
    df_root = df_root.set_index(COLUMN_INDEX)
    for file in files_for_update:
        file_name = f"./Files for update/{file}"
        wb2 = load_workbook(file_name)
        sheets = wb2.sheetnames
        for sheet in sheets[3:]:
            df_update = pd.read_excel(file_name,sheet_name= sheet,index_col=COLUMN_INDEX, header=10)
            df_root.update(df_update)
        df_root=df_root.reset_index()
        df_root=df_root.reindex(df_header_list,axis=1)
        # for the sample i use another file and not the really main data for update
        wb = load_workbook("2.xlsx")
        ws = wb["DATA"]
        for r in dataframe_to_rows(df_root, index=False, header=False):
            ws.append(r)
        create_beauty(df_root.shape[0], len(df_root.columns), ws)
        add_filter(ws)
        wb.save("2.xlsx")
        comments_for_data_update(file_name, "2.xlsx")
        insert_formulas("2.xlsx")

def create_beauty(amount_of_rows,amount_of_columns,worksheet):
    """makes a same cell format. to a data only"""
    for row in range(12,amount_of_rows+13):
        for cell in range(1,amount_of_columns):
            source_cell = worksheet.cell(column=cell, row=12)
            target_cell = worksheet.cell(column=cell, row=row)
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)

            if source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)

            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)
    worksheet.delete_rows(12)


def create_file_director(name,list_of_managers):
    """Creates directory and a file according to a format for director"""
    file_folder = f"Export/{name}"
    pathlib.Path(file_folder).mkdir(parents=True, exist_ok=True)
    wb = load_workbook(FORMATE)
    ws = wb["format"]
    for manager in list_of_managers:
        ws2 = wb.copy_worksheet(ws)
        ws2.title = manager
        ws2.sheet_properties.tabColor = COLORS_LIST[list_of_managers.index(manager)]
        ws2.sheet_view.showGridLines = False
    del wb["format"]
    wb.save(f"{file_folder}/{name}.xlsx")
    return file_folder

def create_file_manager(folder,name):
    """Creates directory and a file according to a format for manager"""
    file_folder = f"{folder}/{name}"
    pathlib.Path(file_folder).mkdir(parents=True, exist_ok=True)
    wb = load_workbook(FORMATE)
    ws = wb["format"]
    ws2 = wb.copy_worksheet(ws)
    ws2.title = name
    ws2.sheet_properties.tabColor = COLORS_LIST[1]
    ws2.sheet_view.showGridLines = False
    del wb["format"]
    wb.save(f"{file_folder}/{name}.xlsx")
    return file_folder

def lock_comment_formulas(target_file, source_file):
    insert_formulas(target_file)
    comments(source_file, target_file)
    lock_sheet(target_file)


create_root_file()
update_main_data(DATA_FILE)

